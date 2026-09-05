"""End-to-end tests: the pipeline on synthetic data, the deliverables, and the app pages."""

from __future__ import annotations

import sys
from datetime import date
from pathlib import Path

import polars as pl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from core import report  # noqa: E402
from core.config import DiagnosticConfig  # noqa: E402
from core.pipeline import run_diagnostic  # noqa: E402

APP = Path(__file__).resolve().parent.parent / "app"


# --------------------------------------------------------------------------- #
# Pipeline behaviour on realistic-but-awkward data
# --------------------------------------------------------------------------- #

def test_pipeline_runs_and_covers_every_sku(result):
    df = result.sku_frame
    # 8 stocked + 1 sold-but-never-stocked
    assert df.height == 9
    assert df["sku"].n_unique() == 9
    assert result.rollup["n_skus"] == 9


def test_padded_sku_codes_reconcile(result):
    """Trailing whitespace in the stock file must not break the join."""
    row = result.sku_frame.filter(pl.col("sku") == "0001").row(0, named=True)
    assert row["description"] == "OVERSTOCKED ITEM"  # joined despite the padding
    assert row["on_hand"] == 5000
    assert row["months_with_sales"] > 0  # matched to sales too


def test_partial_month_excluded_and_forecast_starts_there(result):
    assert result.meta["dropped_partial_period"] == date(2026, 9, 1)
    assert result.meta["last_training_period"] == date(2026, 8, 1)
    assert result.meta["forecast_start"] == date(2026, 9, 1)


def test_overstocked_sku_is_liquidate(result):
    row = result.sku_frame.filter(pl.col("sku") == "0001").row(0, named=True)
    assert row["excess_units"] > 0
    assert row["shortfall_units"] == 0
    assert row["recommendation"].startswith("Liquidate")


def test_understocked_sku_is_buy_and_respects_moq(result):
    row = result.sku_frame.filter(pl.col("sku") == "0002").row(0, named=True)
    assert row["shortfall_units"] > 0
    assert row["buy_qty"] >= 50  # MOQ for this SKU
    assert row["recommendation"] == "Buy"


def test_sold_but_not_stocked_is_zero_on_hand_and_tagged(result):
    row = result.sku_frame.filter(pl.col("sku") == "0009").row(0, named=True)
    assert row["on_hand"] == 0
    assert row["on_hand_source"] == "assumed_zero"


def test_never_sold_stock_is_dead(result):
    row = result.sku_frame.filter(pl.col("sku") == "0006").row(0, named=True)
    assert row["is_dead"]
    assert row["target"] == 0


def test_missing_lead_time_uses_default_and_is_flagged(result):
    row = result.sku_frame.filter(pl.col("sku") == "0003").row(0, named=True)
    assert row["lead_time_source"] == "default"
    assert row["lead_time"] > 0


def test_cost_is_proxied_from_price_when_absent(result):
    row = result.sku_frame.filter(pl.col("sku") == "0001").row(0, named=True)
    assert row["cost_source"] == "price_proxy"
    assert row["unit_cost_effective"] == pytest.approx(2.50)


def test_real_unit_cost_is_used_when_present(dataset):
    cfg = DiagnosticConfig(client_name="Fixture Co", min_history_months=4)
    r = run_diagnostic(
        dataset["sales"], dataset["stocks_with_cost"], dataset["suppliers"], None, cfg,
        today=date(2026, 9, 3),
    )
    row = r.sku_frame.filter(pl.col("sku") == "0001").row(0, named=True)
    assert row["cost_source"] == "actual"
    assert row["unit_cost_effective"] == pytest.approx(1.75)


def test_quality_score_and_checks_present(result):
    assert 0 <= result.quality_score <= 100
    flagged = {r["check"] for r in result.checks.iter_rows(named=True) if r["skus"]}
    assert {"lead_time_defaulted", "cost_proxied", "negative_inventory"} <= flagged


# --------------------------------------------------------------------------- #
# Optional purchase orders
# --------------------------------------------------------------------------- #

def test_without_pos_avoidable_is_na(result):
    assert result.rollup["avoidable_value"] is None
    assert result.rollup["has_purchase_orders"] is False


def test_with_pos_avoidable_appears_and_orphans_are_flagged(result_with_pos):
    ro = result_with_pos.rollup
    assert ro["has_purchase_orders"] is True
    assert ro["avoidable_value"] is not None and ro["avoidable_value"] > 0
    orphan = result_with_pos.checks.filter(pl.col("check") == "orphan_po")
    assert orphan["skus"][0] == 1  # SKU 9999


def test_pos_reduce_shortfall_and_never_double_count(result, result_with_pos):
    """Inbound stock offsets the need; buckets still reconcile exactly."""
    before = result.sku_frame.filter(pl.col("sku") == "0002").row(0, named=True)
    after = result_with_pos.sku_frame.filter(pl.col("sku") == "0002").row(0, named=True)
    assert after["shortfall_units"] < before["shortfall_units"]

    for r in result_with_pos.sku_frame.iter_rows(named=True):
        above = max(r["on_hand"] + r["open_po"] - r["target"], 0)
        assert r["excess_units"] + r["avoidable_units"] == pytest.approx(above)


# --------------------------------------------------------------------------- #
# Deliverables
# --------------------------------------------------------------------------- #

def test_pdf_and_excel_are_written(result, tmp_path):
    pdf = report.write_pdf(result, tmp_path / "d.pdf")
    xlsx = report.write_excel(result, tmp_path / "d.xlsx")
    assert pdf.exists() and pdf.stat().st_size > 1000
    assert xlsx.exists() and xlsx.stat().st_size > 1000

    from openpyxl import load_workbook

    wb = load_workbook(xlsx)
    assert "SKU Actions" in wb.sheetnames
    headers = [c.value for c in wb["SKU Actions"][1]]
    for required in ["SKU", "Description", "Stock", "Target", "Excess", "€ Excess",
                     "Recommendation", "Priority (ABC)"]:
        assert required in headers


def test_pdf_states_na_when_no_po_file(result, tmp_path):
    pytest.importorskip("pypdf")
    from pypdf import PdfReader

    pdf = report.write_pdf(result, tmp_path / "d.pdf")
    text = "\n".join(p.extract_text() or "" for p in PdfReader(str(pdf)).pages)
    assert "No purchase-order file supplied" in text
    assert "Methodology" in text


# --------------------------------------------------------------------------- #
# Streamlit pages render without raising
# --------------------------------------------------------------------------- #

@pytest.mark.parametrize(
    "page", ["Home.py", "pages/1_Data_Quality.py", "pages/2_Results.py", "pages/3_Approve.py"]
)
def test_app_pages_render(page, result):
    from streamlit.testing.v1 import AppTest

    at = AppTest.from_file(str(APP / page), default_timeout=90)
    at.session_state["result"] = result
    at.session_state["config"] = result.config
    at.run()
    assert not at.exception, f"{page} raised: {at.exception}"


def test_pages_guard_when_no_run_yet():
    """Every page except Home must degrade gracefully with no result in session."""
    from streamlit.testing.v1 import AppTest

    for page in ("pages/1_Data_Quality.py", "pages/2_Results.py", "pages/3_Approve.py"):
        at = AppTest.from_file(str(APP / page), default_timeout=60)
        at.run()
        assert not at.exception
        assert any("No diagnostic has been run" in i.value for i in at.info)


# --- Purchase-order arrival timing -------------------------------------------------


def _row(result, sku: str):
    return result.sku_frame.filter(pl.col("sku") == sku).row(0, named=True)


def test_multiple_po_batches_for_one_sku_are_summed(result_po_timing):
    """Three batches on one SKU collapse into a single open_po, none of them lost."""
    r = _row(result_po_timing, "0002")
    assert r["open_po"] == pytest.approx(907.0)  # 3 + 4 + 900


def test_batch_beyond_the_coverage_window_does_not_cancel_a_shortfall(result_po_timing):
    """The 900 units land in 2027 - they are committed spend, but cover nothing now.

    Counting them as available (the old behaviour) would wipe out a real shortfall.
    """
    r = _row(result_po_timing, "0002")
    assert r["open_po_in"] == pytest.approx(7.0)      # 3 overdue + 4 inside
    assert r["open_po_beyond"] == pytest.approx(900.0)

    # Enough inbound exists in total to erase the need, but not in time.
    assert r["on_hand"] + r["open_po"] > r["target"]
    assert r["shortfall_units"] > 0
    assert r["shortfall_units"] == pytest.approx(
        max(r["target"] - (r["on_hand"] + r["open_po_in"]), 0.0)
    )


def test_avoidable_still_counts_every_open_po(result_po_timing):
    """A late batch is cancellable, so it belongs in avoidable even though it covers nothing."""
    r = _row(result_po_timing, "0002")
    expected = max(r["on_hand"] + r["open_po"] - r["target"], 0.0) - max(
        r["on_hand"] - r["target"], 0.0
    )
    assert r["avoidable_units"] == pytest.approx(expected)
    assert r["avoidable_units"] > 0


def test_overdue_pos_are_counted_as_inbound_and_flagged(result_po_timing):
    """Overdue means imminent, not absent - but it is surfaced in the audit."""
    assert result_po_timing.meta["n_overdue_pos"] == 1
    assert result_po_timing.meta["skus_with_overdue_po"] == 1

    check = result_po_timing.checks.filter(pl.col("check") == "overdue_po")
    assert check.height == 1
    assert check["skus"][0] == 1


def test_no_po_file_leaves_the_split_columns_at_zero(result):
    """Without POs there is nothing inbound, so shortfall falls back to on-hand alone."""
    assert result.sku_frame["open_po_in"].sum() == 0.0
    assert result.sku_frame["open_po_beyond"].sum() == 0.0
    assert result.meta["n_overdue_pos"] == 0


# --- Per-month forecast values in the export ---------------------------------------


def test_forecast_export_carries_a_column_per_horizon_month(result):
    """The client can read expected sales per SKU per month, straight from the CSV."""
    export = result.forecast_export()
    months = [c for c in export.columns if c.startswith("forecast_")]

    assert len(months) == result.config.forecast_horizon_months
    assert months == sorted(months)
    # The horizon opens on the month after training - the dropped partial month.
    assert months[0] == result.meta["forecast_start"].strftime("forecast_%Y-%m")

    # Every accuracy column survives, and no SKU is added or lost.
    assert set(result.accuracy.columns).issubset(export.columns)
    assert export.height == result.accuracy.height


def test_forecast_export_months_reconcile_to_the_horizon_total(result):
    """Per-month values are the same numbers the engine forecast, not a re-estimate."""
    export = result.forecast_export()
    months = [c for c in export.columns if c.startswith("forecast_")]
    row = export.filter(pl.col("sku") == "0001").row(0, named=True)

    assert all(row[m] >= 0 for m in months)
    assert sum(row[m] for m in months) == pytest.approx(row["demand_horizon"], rel=1e-3)


# --- Homepage assumptions ----------------------------------------------------------


def _home():
    from streamlit.testing.v1 import AppTest

    at = AppTest.from_file(str(APP / "Home.py"), default_timeout=90)
    at.run()
    assert not at.exception, at.exception
    return at


def test_client_name_is_not_an_assumption():
    """It identifies the run and names the deliverables, so it belongs with the files."""
    at = _home()
    labels = [w.label for w in at.text_input]
    assert "Client name" in labels
    # Currency is fixed at the EUR default rather than asked for on every run.
    assert not any("Currency" in l for l in labels)


def test_review_period_is_a_dropdown_and_the_median_checkbox_is_gone():
    at = _home()
    assert any("Review period" in s.label for s in at.selectbox)
    assert not any("Review period" in s.label for s in at.slider)
    # The checkbox became a typed figure, which is more flexible.
    assert not at.checkbox
    assert any(n.label.startswith("Default lead time") for n in at.number_input)


def test_default_lead_time_label_follows_the_unit_dropdown():
    """A bare number is ambiguous - the label has to name the unit chosen above it."""
    at = _home()
    num = next(n for n in at.number_input if n.label.startswith("Default lead time"))
    assert num.label.endswith("(months)")

    unit = next(s for s in at.selectbox if "Lead-time unit" in s.label)
    unit.select("days").run()

    num = next(n for n in at.number_input if n.label.startswith("Default lead time"))
    assert num.label.endswith("(days)")


def test_a_typed_default_lead_time_is_actually_used(dataset):
    """Deriving a median instead would silently discard what the operator typed."""
    cfg = DiagnosticConfig(
        client_name="Fixture Co",
        min_history_months=4,
        default_lead_time=3.0,
        default_lead_time_from_median=False,
    )
    r = run_diagnostic(
        dataset["sales"], dataset["stocks"], dataset["suppliers"], None, cfg,
        today=date(2026, 9, 3),
    )
    assert r.meta["default_lead_time_used"] == 3.0
    # SKU 0003 has no lead time in the fixture, so it must carry the typed default.
    row = r.sku_frame.filter(pl.col("sku") == "0003").row(0, named=True)
    assert row["lead_time_source"] == "default"
    assert row["lead_time"] == 3.0
